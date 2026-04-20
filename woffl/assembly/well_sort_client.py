"""Well Sort Client

Pulls live Databricks data for the Well Sort tab in Scott's Tools:
- mpu.wells.vw_shut_in (daily downtime log)
- mpu.wells.vw_well_test (allocated and info-only tests)
- mpu.wells.vw_well_header (well universe)

Composes two tables client-side:
- Online Wells: wells not fully shut-in today, with their most-recent test
  (allocated or any) plus 2-month trailing-average outlier flags and
  allocated-vs-info drift.
- Shut-In Wells: wells fully shut-in today with shut-in start date,
  reason, last-online date, and last well test.

"Fully shut in today" = SUM(down_hours) >= FULL_DAY_HOURS_THRESHOLD on
MAX(dtdate), aggregated per well_name across well_bores. The shut-in view
is filtered to **producer** enthids only (via vw_well_header.well_type =
'prod' joined on vw_shut_in.dthid = vw_well_header.enthid) — otherwise
dual-purpose wells like H-31 (producer + water injector) show up as shut-in
whenever the injector side is idle, even if the producer is running.
"""

from __future__ import annotations

import re
from typing import Literal

import numpy as np
import pandas as pd

from woffl.assembly.databricks_client import execute_query

FULL_DAY_HOURS_THRESHOLD = 20.0
TESTS_WINDOW_DAYS = 120
OUTLIER_PCT = 0.25
TRAILING_AVG_DAYS = 60


# ── queries ────────────────────────────────────────────────────────────────

CURRENT_SHUT_IN_QUERY = f"""\
WITH producers AS (
    SELECT enthid
    FROM mpu.wells.vw_well_header
    WHERE field = 'MPU' AND well_type = 'prod'
),
max_date AS (SELECT MAX(dtdate) AS md FROM mpu.wells.vw_shut_in),
daily AS (
    SELECT
        s.well_name,
        s.dtdate,
        SUM(CAST(s.down_hours AS DOUBLE)) AS hrs,
        FIRST(s.down_code) AS down_code,
        FIRST(s.down_reason) AS down_reason,
        FIRST(s.down_notes) AS down_notes
    FROM mpu.wells.vw_shut_in s
    JOIN producers p ON s.dthid = p.enthid
    WHERE s.dtdate >= DATE_SUB((SELECT md FROM max_date), 365)
    GROUP BY s.well_name, s.dtdate
)
SELECT
    d.well_name,
    d.dtdate,
    d.hrs,
    d.down_code,
    d.down_reason,
    d.down_notes
FROM daily d
WHERE d.well_name IN (
    SELECT well_name FROM daily
    WHERE dtdate = (SELECT md FROM max_date)
      AND hrs >= {FULL_DAY_HOURS_THRESHOLD}
)
ORDER BY d.well_name, d.dtdate DESC
"""

ALL_TESTS_QUERY = """\
SELECT
    vwt.well_name,
    vwt.wt_date,
    vwt.allocated,
    vwt.whp,
    vwt.form_oil   AS oil_rate,
    vwt.form_wat   AS fwat_rate,
    vwt.form_gas   AS fgas_rate,
    vwt.form_wc,
    vwt.form_gor   AS fgor,
    vwt.lift_wat   AS lwat_rate,
    vwt.lift_gas   AS lgas_rate,
    vwt.totl_wat   AS twat_rate,
    vwt.totl_gas   AS tgas_rate,
    vwt.totl_wc,
    vwt.totl_gor,
    vwt.esp_amps,
    vwt.esp_hz,
    round(vbdc.bhp_cln_value, 2) AS bhp
FROM mpu.wells.vw_well_test vwt
LEFT JOIN mpu.wells.vw_bhp_daily_clean vbdc
    ON vwt.enthid = vbdc.enthid
    AND to_date(vwt.wt_date) = vbdc.tag_date
WHERE vwt.wt_date >= DATE_SUB(current_date(), {days})
ORDER BY vwt.well_name, vwt.wt_date
"""

WELL_HEADER_QUERY = """\
SELECT well_name
FROM mpu.wells.vw_well_header
WHERE field = 'MPU'
"""

MPU_PRODUCERS_QUERY = """\
SELECT DISTINCT well_name
FROM mpu.wells.vw_well_header
WHERE field = 'MPU' AND well_type = 'prod'
"""

PRODUCER_CATALOG_QUERY = """\
SELECT DISTINCT well_name, well_pad, reservoir
FROM mpu.wells.vw_well_header
WHERE field = 'MPU' AND well_type = 'prod'
"""

# For shut-in / LTSI wells: the absolute-latest test (vw_well_test_recent
# already ranks by recency) plus an average of tests within 90 days of that
# last test, which gives a representative pre-shut-in rate even if the well
# has been down for years.
LAST_TEST_EVER_QUERY = """\
WITH producers AS (
    SELECT enthid FROM mpu.wells.vw_well_header
    WHERE field = 'MPU' AND well_type = 'prod'
),
last_dates AS (
    SELECT t.enthid, MAX(t.wt_date) AS last_date
    FROM mpu.wells.vw_well_test t
    JOIN producers p ON t.enthid = p.enthid
    GROUP BY t.enthid
),
near_window AS (
    SELECT t.enthid,
        AVG(t.form_oil) AS near_avg_oil,
        AVG(t.form_wat) AS near_avg_wat,
        AVG(t.form_gas) AS near_avg_gas,
        COUNT(*) AS n_near_tests
    FROM mpu.wells.vw_well_test t
    JOIN last_dates ld ON ld.enthid = t.enthid
    WHERE t.wt_date BETWEEN DATE_SUB(ld.last_date, 90) AND ld.last_date
    GROUP BY t.enthid
)
SELECT
    r.well_name, r.wt_date AS last_date, r.allocated,
    r.form_oil AS last_oil, r.form_wat AS last_wat, r.form_gas AS last_gas,
    r.lift_wat AS last_lwat, r.lift_gas AS last_lgas,
    r.totl_wat AS last_twat, r.totl_gas AS last_tgas,
    r.form_wc AS last_wc, r.totl_wc AS last_totl_wc,
    r.form_gor AS last_gor, r.totl_gor AS last_totl_gor,
    r.esp_amps AS last_esp_amps, r.esp_hz AS last_esp_hz,
    nw.near_avg_oil, nw.near_avg_wat, nw.near_avg_gas, nw.n_near_tests
FROM mpu.wells.vw_well_test_recent r
JOIN producers p ON p.enthid = r.enthid
LEFT JOIN near_window nw ON nw.enthid = r.enthid
WHERE r.rank_recent = 1
"""

# Tag mapping: production XV = MPU_XZ_<pad_number>2<well_number:02d>
#              power-fluid XV = MPU_XZ_<pad_number>4<well_number:02d>
# pad_number comes from mpu.wells.pad_xref / vw_bhp_tags.
# Fetch current open/shut state for every tag matching either pattern.
XV_STATUS_QUERY = """\
WITH producers AS (
    SELECT h.well_name, t.pad_number, t.well_number
    FROM mpu.wells.vw_well_header h
    JOIN mpu.wells.vw_bhp_tags t ON h.enthid = t.enthid
    WHERE h.field = 'MPU' AND h.well_type = 'prod'
),
tag_map AS (
    SELECT well_name, pad_number, well_number,
        CONCAT('MPU_XZ_', LPAD(CAST(pad_number AS STRING), 2, '0'),
               '2', LPAD(CAST(well_number AS STRING), 2, '0')) AS prod_tag,
        CONCAT('MPU_XZ_', LPAD(CAST(pad_number AS STRING), 2, '0'),
               '4', LPAD(CAST(well_number AS STRING), 2, '0')) AS pf_tag
    FROM producers
),
latest AS (
    SELECT Tag, Value, MeasureTime,
        ROW_NUMBER() OVER (PARTITION BY Tag ORDER BY MeasureTime DESC) AS rn
    FROM historian.mpu.measurements_silver
    WHERE MeasureDate >= DATE_SUB(current_date(), 2)
      AND (
        Tag IN (SELECT prod_tag FROM tag_map)
        OR Tag IN (SELECT pf_tag FROM tag_map)
      )
)
SELECT tm.well_name, tm.prod_tag, tm.pf_tag,
    pl.Value AS prod_value, pl.MeasureTime AS prod_time,
    fl.Value AS pf_value, fl.MeasureTime AS pf_time
FROM tag_map tm
LEFT JOIN (SELECT * FROM latest WHERE rn = 1) pl ON pl.Tag = tm.prod_tag
LEFT JOIN (SELECT * FROM latest WHERE rn = 1) fl ON fl.Tag = tm.pf_tag
"""


# ── helpers ────────────────────────────────────────────────────────────────

def _normalize_well_name(name: str) -> str:
    """Normalize well names to MPx-## (e.g., 'B-028' -> 'MPB-28').

    Mirrors well_test_client._normalize_well_name so downstream code sees
    consistent names across GUI pages.
    """
    if not isinstance(name, str):
        return name
    match = re.search(r"(\w+-\d+)", name)
    if not match:
        return name
    well = match.group(1)
    well = re.sub(r"-(0)(?=\d+)", "-", well)
    if not well.startswith("MP"):
        well = "MP" + well
    return well


# ── fetchers ───────────────────────────────────────────────────────────────

def fetch_current_shut_in_history() -> pd.DataFrame:
    """All daily shut-in rows for wells currently fully shut-in.

    Returns columns: well, dtdate (datetime64[ns]), hrs, down_code,
    down_reason, down_notes. One row per well per day in the last 365 days.
    """
    df = execute_query(CURRENT_SHUT_IN_QUERY)
    if df.empty:
        return df
    df = df.rename(columns={"well_name": "well"})
    df["well"] = df["well"].apply(_normalize_well_name)
    df["dtdate"] = pd.to_datetime(df["dtdate"])
    df["hrs"] = pd.to_numeric(df["hrs"], errors="coerce")
    return df


def fetch_xv_status() -> pd.DataFrame:
    """Current production + power-fluid safety-valve status per producer.

    Returns columns: well (MP-prefixed), prod_tag, pf_tag, prod_value,
    prod_time, pf_value, pf_time. Value=1 means the safety valve is open;
    Value=0 means closed.

    Caveat: SSSV open != well flowing. A closed XV reliably indicates the
    well is not producing; an open XV is necessary but not sufficient.
    """
    df = execute_query(XV_STATUS_QUERY)
    if df.empty:
        return df
    df = df.rename(columns={"well_name": "well"})
    df["well"] = df["well"].apply(_normalize_well_name)
    for c in ("prod_value", "pf_value"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    for c in ("prod_time", "pf_time"):
        df[c] = pd.to_datetime(df[c], utc=True, errors="coerce").dt.tz_localize(None)
    return df


def fetch_mpu_producers() -> list[str]:
    """MP-prefixed well names for every MPU producer in the header view."""
    df = execute_query(MPU_PRODUCERS_QUERY)
    if df.empty:
        return []
    return sorted({_normalize_well_name(n) for n in df["well_name"]})


def fetch_recent_tests(days: int = TESTS_WINDOW_DAYS) -> pd.DataFrame:
    """All well tests (allocated and info-only) in the last `days` days."""
    df = execute_query(ALL_TESTS_QUERY.format(days=int(days)))
    if df.empty:
        return df

    df = df.rename(
        columns={
            "well_name": "well",
            "wt_date": "WtDate",
            "bhp": "BHP",
            "oil_rate": "WtOilVol",
            "fwat_rate": "WtWaterVol",
            "fgas_rate": "WtGasVol",
        }
    )
    df["well"] = df["well"].apply(_normalize_well_name)
    df["WtDate"] = pd.to_datetime(df["WtDate"], utc=True).dt.tz_localize(None)
    for col in ["BHP", "WtOilVol", "WtWaterVol", "WtGasVol",
                "whp", "form_wc", "fgor",
                "lwat_rate", "lgas_rate",
                "twat_rate", "tgas_rate", "totl_wc", "totl_gor",
                "esp_amps", "esp_hz"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["allocated"] = df["allocated"].astype(bool)
    df["WtTotalFluid"] = df["WtOilVol"].fillna(0) + df["WtWaterVol"].fillna(0)
    return df.sort_values(["well", "WtDate"]).reset_index(drop=True)


def fetch_last_tests_ever() -> pd.DataFrame:
    """One row per producer: absolute-latest test + 90-day pre-shut-in avg.

    Returns columns keyed by normalized well name:
      well, last_date, last_allocated, last_oil, last_wat, last_gas,
      last_lwat, last_lgas, last_twat, last_tgas, last_wc, last_totl_wc,
      last_gor, last_totl_gor, last_esp_amps, last_esp_hz,
      near_avg_oil, near_avg_wat, near_avg_gas, n_near_tests.
    """
    df = execute_query(LAST_TEST_EVER_QUERY)
    if df.empty:
        return df
    df = df.rename(columns={"well_name": "well", "allocated": "last_allocated"})
    df["well"] = df["well"].apply(_normalize_well_name)
    df["last_date"] = pd.to_datetime(df["last_date"], utc=True, errors="coerce").dt.tz_localize(None)
    numeric_cols = ["last_oil", "last_wat", "last_gas", "last_lwat", "last_lgas",
                    "last_twat", "last_tgas", "last_wc", "last_totl_wc",
                    "last_gor", "last_totl_gor", "last_esp_amps", "last_esp_hz",
                    "near_avg_oil", "near_avg_wat", "near_avg_gas", "n_near_tests"]
    for c in numeric_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def fetch_producer_catalog() -> pd.DataFrame:
    """Producer metadata: one row per producer with well_pad + reservoir."""
    df = execute_query(PRODUCER_CATALOG_QUERY)
    if df.empty:
        return df
    df = df.rename(columns={"well_name": "well"})
    df["well"] = df["well"].apply(_normalize_well_name)
    # A well can appear multiple times if it has multiple bores; keep first
    # non-null reservoir.
    return (df.sort_values("reservoir", na_position="last")
              .drop_duplicates("well", keep="first")
              .reset_index(drop=True))


def derive_lift_type(row: pd.Series) -> str:
    """Infer lift type from the latest test row.

    Rules (first match wins):
      esp_amps > 0        -> esp
      lgas_rate > 0       -> gas_lift
      lwat_rate > 0       -> jet_pump
      else                -> free_flow
    """
    if pd.notna(row.get("esp_amps")) and row["esp_amps"] > 0:
        return "esp"
    if pd.notna(row.get("lgas_rate")) and row["lgas_rate"] > 0:
        return "gas_lift"
    if pd.notna(row.get("lwat_rate")) and row["lwat_rate"] > 0:
        return "jet_pump"
    return "free_flow"


# ── composition ────────────────────────────────────────────────────────────

def _latest(group: pd.DataFrame) -> pd.Series | None:
    if group.empty:
        return None
    return group.iloc[-1]


def _pct_diff(a: float, b: float) -> float:
    """Percent difference of a vs b, using b as baseline. NaN-safe."""
    if pd.isna(a) or pd.isna(b) or b == 0:
        return np.nan
    return (a - b) / b


def _trailing_avg_excluding(
    well_tests: pd.DataFrame,
    display_date: pd.Timestamp,
    days: int,
    col: str,
) -> float:
    start = display_date - pd.Timedelta(days=days)
    mask = (
        (well_tests["WtDate"] >= start)
        & (well_tests["WtDate"] < display_date)
    )
    values = well_tests.loc[mask, col].dropna()
    if values.empty:
        return np.nan
    return float(values.mean())


def _xv_lookup(xv_df: pd.DataFrame) -> dict[str, dict]:
    """Flatten XV dataframe into a dict keyed by normalized well name."""
    if xv_df is None or xv_df.empty:
        return {}
    return xv_df.set_index("well").to_dict("index")


def classify_wells(
    producers: list[str],
    shut_in_df: pd.DataFrame,
    xv_df: pd.DataFrame | None = None,
    trust_xv: bool = True,
) -> tuple[set[str], set[str]]:
    """Split producers into (online_wells, shut_in_wells).

    Rules (asymmetric — XV can only rescue wells OUT of the shut-in list):
    - Not in vw_shut_in -> online. XV is ignored here; edge cases like
      temporary flowback (H-31) would otherwise get misclassified.
    - In vw_shut_in + ProdXV = 1 (trust_xv=True) -> online ("just restarted";
      the daily log is stale).
    - In vw_shut_in + ProdXV != 1 -> shut-in.
    """
    shut_log = set(shut_in_df["well"].unique()) if not shut_in_df.empty else set()
    xv_map = _xv_lookup(xv_df) if trust_xv else {}

    online, shut = set(), set()
    for well in producers:
        if well not in shut_log:
            online.add(well)
            continue
        x = xv_map.get(well)
        prod = x.get("prod_value") if x else None
        if prod == 1:
            online.add(well)  # log lags a restart
        else:
            shut.add(well)
    return online, shut


def build_online_table(
    tests_df: pd.DataFrame,
    shut_in_df: pd.DataFrame,
    producers: list[str],
    mode: Literal["allocated", "any"] = "allocated",
    stale_days: int = 60,
    xv_df: pd.DataFrame | None = None,
    online_wells: set[str] | None = None,
    catalog_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Compose the Online Wells table.

    Online = every MPU producer NOT in the currently-shut-in set. A producer
    with no recent test still appears (with NaN rates), flagged via
    `StaleTest = DaysSinceTest > stale_days`.

    Display test is picked per `mode`:
    - "allocated": most recent allocated test of any age. Falls back to
      latest info-only when no allocated test exists.
    - "any": most recent test regardless of flag.

    Also attaches latest-alloc and latest-info dates plus 2-month trailing
    deviations so outliers/stale tests can be flagged separately.
    """
    if online_wells is None:
        online_wells, _ = classify_wells(producers, shut_in_df, xv_df)
    online_list = sorted(online_wells)
    shut_log = set(shut_in_df["well"].unique()) if not shut_in_df.empty else set()
    today = pd.Timestamp.now().normalize()
    xv_map = _xv_lookup(xv_df)
    cat_map: dict[str, dict] = (
        catalog_df.set_index("well").to_dict("index")
        if catalog_df is not None and not catalog_df.empty else {}
    )

    def _xv_cols(well: str) -> dict:
        x = xv_map.get(well)
        if x is None:
            return {"ProdXV": np.nan, "PFXV": np.nan, "XVTime": pd.NaT,
                    "JustRestarted": False}
        return {
            "ProdXV": x.get("prod_value"),
            "PFXV": x.get("pf_value"),
            "XVTime": x.get("prod_time") if pd.notna(x.get("prod_time")) else x.get("pf_time"),
            # Daily log still has well as shut-in but XV says it's flowing —
            # the well just came back online since the last vw_shut_in refresh.
            "JustRestarted": x.get("prod_value") == 1 and well in shut_log,
        }

    rows = []
    for well in online_list:
        wt = tests_df[tests_df["well"] == well]

        cat = cat_map.get(well, {})
        if wt.empty:
            row = {
                "Well": well,
                "Pad": cat.get("well_pad"),
                "Reservoir": cat.get("reservoir"),
                "LiftType": None,
                "PopsPad": False,
                "TestDate": pd.NaT, "DaysSinceTest": np.nan,
                "StaleTest": True, "Allocated": False, "FallbackUsed": True,
                "Oil": np.nan, "Water": np.nan, "Gas": np.nan,
                "LiftWater": np.nan, "LiftGas": np.nan,
                "TotalWater": np.nan, "TotalGas": np.nan,
                "WC": np.nan, "TotalWC": np.nan,
                "GOR": np.nan, "TotalGOR": np.nan,
                "BHP": np.nan, "WHP": np.nan,
                "Oil_2moAvg": np.nan, "Wat_2moAvg": np.nan,
                "OilDev": np.nan, "WatDev": np.nan, "FlagOutlier": False,
                "AllocVsInfoOilPct": np.nan,
                "LatestAllocDate": pd.NaT, "LatestInfoDate": pd.NaT,
            }
            row.update(_xv_cols(well))
            rows.append(row)
            continue

        latest_any = wt.iloc[-1]
        alloc = wt[wt["allocated"]]
        info = wt[~wt["allocated"]]
        latest_alloc = alloc.iloc[-1] if not alloc.empty else None
        latest_info = info.iloc[-1] if not info.empty else None

        if mode == "allocated":
            display = latest_alloc if latest_alloc is not None else latest_any
            fallback_used = latest_alloc is None
        else:
            display = latest_any
            fallback_used = False

        days_since = (today - display["WtDate"].normalize()).days

        oil_avg = _trailing_avg_excluding(wt, display["WtDate"],
                                          TRAILING_AVG_DAYS, "WtOilVol")
        wat_avg = _trailing_avg_excluding(wt, display["WtDate"],
                                          TRAILING_AVG_DAYS, "WtWaterVol")
        oil_dev = _pct_diff(display["WtOilVol"], oil_avg)
        wat_dev = _pct_diff(display["WtWaterVol"], wat_avg)
        flag_outlier = (
            (not pd.isna(oil_dev) and abs(oil_dev) > OUTLIER_PCT)
            or (not pd.isna(wat_dev) and abs(wat_dev) > OUTLIER_PCT)
        )

        alloc_vs_info_oil = np.nan
        if latest_alloc is not None and latest_info is not None:
            alloc_vs_info_oil = _pct_diff(
                latest_info["WtOilVol"], latest_alloc["WtOilVol"]
            )

        row = {
            "Well": well,
            "Pad": cat.get("well_pad"),
            "Reservoir": cat.get("reservoir"),
            "LiftType": derive_lift_type(display),
            "PopsPad": False,
            "TestDate": display["WtDate"],
            "DaysSinceTest": days_since,
            "StaleTest": days_since > stale_days,
            "Allocated": bool(display["allocated"]),
            "FallbackUsed": fallback_used,
            "Oil": display["WtOilVol"],
            "Water": display["WtWaterVol"],
            "Gas": display["WtGasVol"],
            "LiftWater": display.get("lwat_rate", np.nan),
            "LiftGas": display.get("lgas_rate", np.nan),
            "TotalWater": display.get("twat_rate", np.nan),
            "TotalGas": display.get("tgas_rate", np.nan),
            "WC": display.get("form_wc", np.nan),
            "TotalWC": display.get("totl_wc", np.nan),
            "GOR": display.get("fgor", np.nan),
            "TotalGOR": display.get("totl_gor", np.nan),
            "BHP": display.get("BHP", np.nan),
            "WHP": display.get("whp", np.nan),
            "Oil_2moAvg": oil_avg,
            "Wat_2moAvg": wat_avg,
            "OilDev": oil_dev,
            "WatDev": wat_dev,
            "FlagOutlier": flag_outlier,
            "AllocVsInfoOilPct": alloc_vs_info_oil,
            "LatestAllocDate": latest_alloc["WtDate"] if latest_alloc is not None else pd.NaT,
            "LatestInfoDate": latest_info["WtDate"] if latest_info is not None else pd.NaT,
        }
        row.update(_xv_cols(well))
        rows.append(row)

    return pd.DataFrame(rows)


def build_shut_in_table(
    shut_in_df: pd.DataFrame,
    tests_df: pd.DataFrame,
    xv_df: pd.DataFrame | None = None,
    shut_in_wells: set[str] | None = None,
    catalog_df: pd.DataFrame | None = None,
    last_tests_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Compose the Shut-In Wells table.

    Uses the daily shut-in history (`shut_in_df`) to compute:
    - ShutInSince: earliest consecutive full-day-down date ending at max_date.
    - LastOnlineDate: last date where hrs < FULL_DAY_HOURS_THRESHOLD.
    - CurrentCode/CurrentReason: taken from the most recent row.

    Joins latest well test date/rates from `tests_df`.
    """
    xv_map = _xv_lookup(xv_df)
    cat_map: dict[str, dict] = (
        catalog_df.set_index("well").to_dict("index")
        if catalog_df is not None and not catalog_df.empty else {}
    )
    last_map: dict[str, dict] = (
        last_tests_df.set_index("well").to_dict("index")
        if last_tests_df is not None and not last_tests_df.empty else {}
    )
    if shut_in_df.empty and not shut_in_wells:
        return pd.DataFrame()

    max_date = shut_in_df["dtdate"].max() if not shut_in_df.empty else pd.NaT

    # If caller gave an explicit shut-in set (XV-based), iterate over that;
    # otherwise fall back to every well in the vw_shut_in log.
    if shut_in_wells is None:
        shut_in_wells = set(shut_in_df["well"].unique())

    rows = []
    for well in sorted(shut_in_wells):
        grp = shut_in_df[shut_in_df["well"] == well]
        grp = grp.sort_values("dtdate", ascending=False).reset_index(drop=True)

        cat = cat_map.get(well, {})
        last = last_map.get(well)  # absolute-latest test, any age

        def _bench_cols():
            """Rates for the shut-in row.

            Prefer the all-time-latest test (`last_tests_df`) since LTSI wells
            may have been down for years and won't show up in the 180-day
            tests_df. Falls back to latest within tests_df if no all-time row
            exists, and finally to NaN.
            """
            if last is not None and pd.notna(last.get("last_date")):
                # Build a synthetic "row" so derive_lift_type works.
                synthetic = pd.Series({
                    "esp_amps": last.get("last_esp_amps"),
                    "lgas_rate": last.get("last_lgas"),
                    "lwat_rate": last.get("last_lwat"),
                })
                return {
                    "LiftType": derive_lift_type(synthetic),
                    "Oil": last.get("last_oil"),
                    "Water": last.get("last_wat"),
                    "Gas": last.get("last_gas"),
                    "LiftWater": last.get("last_lwat"),
                    "LiftGas": last.get("last_lgas"),
                    "TotalWater": last.get("last_twat"),
                    "TotalGas": last.get("last_tgas"),
                    "WC": last.get("last_wc"),
                    "TotalWC": last.get("last_totl_wc"),
                    "GOR": last.get("last_gor"),
                    "TotalGOR": last.get("last_totl_gor"),
                    "NearAvgOil": last.get("near_avg_oil"),
                    "NearAvgWater": last.get("near_avg_wat"),
                    "NearAvgGas": last.get("near_avg_gas"),
                    "NTestsNear": last.get("n_near_tests"),
                }
            # Fallback: use whatever's in tests_df
            wt_all = tests_df[tests_df["well"] == well] if not tests_df.empty else pd.DataFrame()
            if wt_all.empty:
                return {
                    "LiftType": None, "Oil": np.nan, "Water": np.nan, "Gas": np.nan,
                    "LiftWater": np.nan, "LiftGas": np.nan,
                    "TotalWater": np.nan, "TotalGas": np.nan,
                    "WC": np.nan, "TotalWC": np.nan,
                    "GOR": np.nan, "TotalGOR": np.nan,
                    "NearAvgOil": np.nan, "NearAvgWater": np.nan,
                    "NearAvgGas": np.nan, "NTestsNear": np.nan,
                }
            r = wt_all.iloc[-1]
            return {
                "LiftType": derive_lift_type(r),
                "Oil": r["WtOilVol"], "Water": r["WtWaterVol"], "Gas": r["WtGasVol"],
                "LiftWater": r.get("lwat_rate", np.nan),
                "LiftGas": r.get("lgas_rate", np.nan),
                "TotalWater": r.get("twat_rate", np.nan),
                "TotalGas": r.get("tgas_rate", np.nan),
                "WC": r.get("form_wc", np.nan),
                "TotalWC": r.get("totl_wc", np.nan),
                "GOR": r.get("fgor", np.nan),
                "TotalGOR": r.get("totl_gor", np.nan),
                "NearAvgOil": np.nan, "NearAvgWater": np.nan,
                "NearAvgGas": np.nan, "NTestsNear": np.nan,
            }

        # Under the current classify_wells rules, every shut-in well is in
        # vw_shut_in. Skip defensively if that ever isn't the case.
        if grp.empty:
            continue

        streak_start = max_date
        expected = max_date
        for _, r in grp.iterrows():
            if r["dtdate"] == expected and r["hrs"] >= FULL_DAY_HOURS_THRESHOLD:
                streak_start = r["dtdate"]
                expected = expected - pd.Timedelta(days=1)
            else:
                break

        online_mask = grp["hrs"] < FULL_DAY_HOURS_THRESHOLD
        last_online = grp.loc[online_mask, "dtdate"].max() if online_mask.any() else pd.NaT

        current = grp.iloc[0]

        # Prefer the all-time-latest test (for LTSI wells shut in > 180 days).
        if last is not None and pd.notna(last.get("last_date")):
            last_test_date = last["last_date"]
            last_oil = last.get("last_oil")
            last_wat = last.get("last_wat")
            last_gas = last.get("last_gas")
        else:
            wt = tests_df[tests_df["well"] == well]
            if not wt.empty:
                last_test = wt.iloc[-1]
                last_test_date = last_test["WtDate"]
                last_oil = last_test["WtOilVol"]
                last_wat = last_test["WtWaterVol"]
                last_gas = last_test["WtGasVol"]
            else:
                last_test_date = pd.NaT
                last_oil = last_wat = last_gas = np.nan

        x = xv_map.get(well, {})
        prod_xv = x.get("prod_value") if x else np.nan
        row = {
            "Well": well,
            "Pad": cat.get("well_pad"),
            "Reservoir": cat.get("reservoir"),
            "PopsPad": False,
            "ShutInSince": streak_start,
            "CurrentCode": current["down_code"],
            "CurrentReason": current["down_reason"],
            "Notes": current["down_notes"],
            "DownHours": current["hrs"],
            "LastOnlineDate": last_online,
            "LastTestDate": last_test_date,
            "LastTestOil": last_oil,
            "LastTestWater": last_wat,
            "LastTestGas": last_gas,
            "ProdXV": prod_xv,
            "PFXV": x.get("pf_value") if x else np.nan,
            "XVTime": (x.get("prod_time") if x and pd.notna(x.get("prod_time"))
                       else x.get("pf_time") if x else pd.NaT),
        }
        row.update(_bench_cols())
        rows.append(row)

    return pd.DataFrame(rows).sort_values("ShutInSince", ascending=False).reset_index(drop=True)


# ── bench helpers ──────────────────────────────────────────────────────────

LTSI_CODE_PREFIX = "T"  # T01 mech, T02 reservoir, T03 convert, T05 P&A


def split_offline_ltsi(
    shut_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split a shut-in dataframe into (offline, ltsi).

    LTSI = CurrentCode starts with 'T' (long-term shut-in codes). Offline =
    everything else (SI, D-codes, F-codes, etc.).
    """
    if shut_df.empty:
        return shut_df, shut_df
    code = shut_df["CurrentCode"].fillna("").astype(str)
    is_ltsi = code.str.startswith(LTSI_CODE_PREFIX)
    return (
        shut_df[~is_ltsi].reset_index(drop=True),
        shut_df[is_ltsi].reset_index(drop=True),
    )


def apply_pops_pad(
    df: pd.DataFrame,
    pads_with_separation: set[str],
    overrides: dict[str, bool] | None = None,
) -> pd.DataFrame:
    """Set the PopsPad column based on pad-level flags + per-well overrides.

    pads_with_separation: set of pad letters (e.g. {'E', 'S'}) that have on-pad
      production separation. Every well on those pads gets PopsPad=True.
    overrides: per-well boolean overrides (e.g. {'MPS-08': True}) applied last,
      so individual wells can deviate from their pad default.
    """
    if df.empty or "Pad" not in df.columns:
        return df
    df = df.copy()
    df["PopsPad"] = df["Pad"].isin(pads_with_separation).fillna(False)
    if overrides:
        for well, val in overrides.items():
            df.loc[df["Well"] == well, "PopsPad"] = bool(val)
    return df


_CUMULATIVE_COLS: dict[str, str] = {
    "Oil": "cum_oil",
    "Water": "cum_fwat",
    "Gas": "cum_fgas",
    "TotalWater": "cum_twat",
    "TotalGas": "cum_tgas",
}


def add_cumulative_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add running-sum columns (cum_oil, cum_fwat, cum_fgas, cum_twat, cum_tgas).

    Computed in the df's current row order — sort before calling if a specific
    cumulation order is wanted.
    """
    if df.empty:
        return df
    df = df.copy()
    for src, dst in _CUMULATIVE_COLS.items():
        if src in df.columns:
            df[dst] = df[src].fillna(0).cumsum()
    return df


# Bench-sheet column order (matches MPU_Well_Bench spreadsheet exactly).
_BENCH_BASE_COLS = [
    "Well", "Pad", "Reservoir",
    "Oil", "Water", "Gas",
    "LiftWater", "LiftGas",
    "TotalWater", "TotalGas",
    "WC", "TotalWC",
    "GOR", "TotalGOR",
    "LiftType", "PopsPad",
]
_BENCH_CUM_COLS = ["cum_oil", "cum_fwat", "cum_fgas", "cum_twat", "cum_tgas"]
_BENCH_OFFLINE_EXTRA = [
    "LastTestDate", "NearAvgOil", "NearAvgWater", "NearAvgGas", "NTestsNear",
    "CurrentReason", "Notes", "DownHours", "CurrentCode",
]

_BENCH_RENAME = {
    "Well": "Well", "Pad": "Pad", "Reservoir": "Reservoir",
    "Oil": "Oil", "Water": "Form Water", "Gas": "Form Gas",
    "LiftWater": "Power Fluid", "LiftGas": "Lift Gas",
    "TotalWater": "Total Water", "TotalGas": "Total Gas",
    "WC": "Form WC", "TotalWC": "Total WC",
    "GOR": "Form GOR", "TotalGOR": "Total GOR",
    "LiftType": "Lift Type", "PopsPad": "Pops Pad",
    "CurrentReason": "Down Reason", "Notes": "Down Note",
    "DownHours": "Down Hrs", "CurrentCode": "Down Code",
    "LastTestDate": "Last Test Date",
    "NearAvgOil": "Near Avg Oil", "NearAvgWater": "Near Avg Water",
    "NearAvgGas": "Near Avg Gas", "NTestsNear": "# Near Tests",
    "cum_oil": "C_Oil", "cum_fwat": "C_Form Water", "cum_fgas": "C_Form Gas",
    "cum_twat": "C_Total Water", "cum_tgas": "C_Total Gas",
}

# Bench columns that MUST be numeric when Excel opens (prevents the
# "convert to number" prompt). Names match post-rename labels.
_BENCH_NUMERIC_COLS = {
    "Oil", "Form Water", "Form Gas", "Power Fluid", "Lift Gas",
    "Total Water", "Total Gas", "Form WC", "Total WC", "Form GOR", "Total GOR",
    "C_Oil", "C_Form Water", "C_Form Gas", "C_Total Water", "C_Total Gas",
    "Down Hrs",
    "Near Avg Oil", "Near Avg Water", "Near Avg Gas", "# Near Tests",
}


def _bench_frame(df: pd.DataFrame, include_offline_extras: bool) -> pd.DataFrame:
    """Select + rename bench columns, add cumulative sums, in spreadsheet order.

    Cumulative columns are only emitted for the online sheet (matches the
    reference workbook — cum numbers on shut-in wells are meaningless).
    """
    if df.empty:
        return df
    if not include_offline_extras:
        df = add_cumulative_columns(df)
    cols = list(_BENCH_BASE_COLS)
    if include_offline_extras:
        cols += [c for c in _BENCH_OFFLINE_EXTRA if c in df.columns]
    else:
        cols += [c for c in _BENCH_CUM_COLS if c in df.columns]
    existing = [c for c in cols if c in df.columns]
    out = df[existing].rename(columns=_BENCH_RENAME)
    # Percent conversion (our source is a fraction; bench sheet shows 0-100).
    for c in ("Form WC", "Total WC"):
        if c in out.columns:
            out[c] = out[c] * 100
    # Force numeric dtype on all rate/cum columns so openpyxl writes them
    # as real numbers (Excel won't prompt "Convert to Number").
    for c in out.columns:
        if c in _BENCH_NUMERIC_COLS:
            out[c] = pd.to_numeric(out[c], errors="coerce")
    return out


def export_bench_xlsx(
    online_df: pd.DataFrame,
    offline_df: pd.DataFrame,
    ltsi_df: pd.DataFrame,
) -> bytes:
    """Build a 3-sheet bench workbook matching the MPU_Well_Bench format.

    - Numeric columns are written as real numbers (no "convert to number"
      prompt in Excel).
    - Each sheet gets an auto-filter applied to the header row for easy
      sorting/filtering.
    - Rate cells get a thousands-separator display format; WC columns get
      one decimal.
    """
    from io import BytesIO
    from openpyxl.utils import get_column_letter

    frames = {
        "online":  _bench_frame(online_df, False),
        "offline": _bench_frame(offline_df, True),
        "ltsi":    _bench_frame(ltsi_df, True),
    }

    # Per-column Excel display format (applied post-write).
    int_cols = {
        "Oil", "Form Water", "Form Gas", "Power Fluid", "Lift Gas",
        "Total Water", "Total Gas", "Form GOR", "Total GOR",
        "C_Oil", "C_Form Water", "C_Form Gas", "C_Total Water", "C_Total Gas",
        "Near Avg Oil", "Near Avg Water", "Near Avg Gas", "# Near Tests",
    }
    one_decimal_cols = {"Form WC", "Total WC", "Down Hrs"}

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, frame in frames.items():
            frame.to_excel(xw, sheet_name=name, index=False)
            ws = xw.book[name]
            if ws.max_row < 1 or ws.max_column < 1:
                continue
            # Auto-filter on header row covering the whole data range.
            ws.auto_filter.ref = ws.dimensions
            # Freeze header for easy scrolling.
            ws.freeze_panes = "A2"
            # Apply number formats.
            headers = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(headers, start=1):
                if header is None:
                    continue
                if header in int_cols:
                    fmt = "#,##0"
                elif header in one_decimal_cols:
                    fmt = "0.0"
                else:
                    continue
                letter = get_column_letter(col_idx)
                for cell in ws[letter][1:]:  # skip header
                    cell.number_format = fmt
    return buf.getvalue()
